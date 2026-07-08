"""Load unstructured data into the vector store.

Supported inputs:

* **Files** — ``.txt``, ``.md``, ``.pdf``, ``.csv``, ``.xlsx`` (see :data:`LOADERS`).
* **URLs** — any web page; HTML is fetched and reduced to readable text.

Everything funnels through :func:`chunk_document`, so the rest of the pipeline
never sees a file or a URL — only :class:`Chunk` objects. Each chunk can be
tagged with a ``tenant_id`` so multiple users/customers share one store without
seeing each other's data.
"""

from __future__ import annotations

import csv
from collections.abc import Callable, Iterable
from pathlib import Path
from urllib.parse import urldefrag, urljoin, urlparse

from .chunking import chunk_document
from .config import Settings, settings
from .embeddings import Embedder, build_embedder
from .types import Chunk
from .vector_store import VectorStore

TEXT_SUFFIXES = {".txt", ".md", ".markdown", ".rst"}


def _load_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def _load_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - import guard
        raise ImportError("pypdf is required to ingest PDF files.") from exc

    reader = PdfReader(str(path))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(pages)


def _rows_to_text(rows: Iterable[Iterable[object]]) -> str:
    """Render tabular rows as pipe-delimited lines (one row per line)."""
    lines = []
    for row in rows:
        cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if cells:
            lines.append(" | ".join(cells))
    return "\n".join(lines)


def _load_csv(path: Path) -> str:
    with path.open(newline="", encoding="utf-8", errors="ignore") as fh:
        return _rows_to_text(csv.reader(fh))


def _load_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - import guard
        raise ImportError("openpyxl is required to ingest .xlsx files.") from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        body = _rows_to_text(ws.iter_rows(values_only=True))
        if body:
            parts.append(f"# Sheet: {ws.title}\n{body}")
    wb.close()
    return "\n\n".join(parts)


LOADERS: dict[str, Callable[[Path], str]] = {
    **{suffix: _load_text for suffix in TEXT_SUFFIXES},
    ".pdf": _load_pdf,
    ".csv": _load_csv,
    ".xlsx": _load_xlsx,
}


def discover_files(paths: Iterable[str | Path]) -> list[Path]:
    """Expand files and directories into a flat, sorted list of loadable files."""
    found: list[Path] = []
    for raw in paths:
        p = Path(raw)
        if p.is_dir():
            found.extend(
                f for f in sorted(p.rglob("*")) if f.suffix.lower() in LOADERS
            )
        elif p.is_file() and p.suffix.lower() in LOADERS:
            found.append(p)
    return found


_USER_AGENT = "NextCX-RAG/1.0"


def _fetch_html(url: str, *, timeout: int = 20) -> str:
    """GET a URL and return its raw HTML (raises on network/HTTP error)."""
    try:
        import requests
    except ImportError as exc:  # pragma: no cover - import guard
        raise ImportError("requests is required to ingest URLs.") from exc

    resp = requests.get(url, timeout=timeout, headers={"User-Agent": _USER_AGENT})
    resp.raise_for_status()
    return resp.text


def _html_to_text(html: str) -> str:
    """Strip boilerplate tags and return readable text."""
    try:
        from bs4 import BeautifulSoup
    except ImportError as exc:  # pragma: no cover - import guard
        raise ImportError("beautifulsoup4 is required to ingest URLs.") from exc

    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript", "header", "footer", "nav"]):
        tag.decompose()
    return soup.get_text(separator="\n", strip=True)


def _extract_links(html: str, base_url: str) -> list[str]:
    """Absolute, de-fragmented http(s) links found in ``html``."""
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(html, "html.parser")
    links: list[str] = []
    for a in soup.find_all("a", href=True):
        absolute = urldefrag(urljoin(base_url, a["href"]))[0]
        if absolute.startswith(("http://", "https://")):
            links.append(absolute)
    return links


def fetch_url(url: str, *, timeout: int = 20) -> str:
    """Fetch a single web page and reduce it to readable text."""
    return _html_to_text(_fetch_html(url, timeout=timeout))


def crawl_site(
    start_url: str,
    *,
    max_pages: int = 20,
    same_domain: bool = True,
    timeout: int = 20,
    fetcher: Callable[[str], str] | None = None,
) -> list[tuple[str, str]]:
    """Breadth-first crawl from ``start_url``, returning ``(url, text)`` pairs.

    Stays on the start URL's domain by default and stops after ``max_pages``.
    ``fetcher`` (url -> html) is injectable for testing; it defaults to a real
    HTTP GET. Pages that fail to fetch are skipped, not fatal.
    """
    fetch = fetcher or (lambda u: _fetch_html(u, timeout=timeout))
    domain = urlparse(start_url).netloc
    queue: list[str] = [urldefrag(start_url)[0]]
    visited: set[str] = set()
    pages: list[tuple[str, str]] = []

    while queue and len(pages) < max_pages:
        url = queue.pop(0)
        if url in visited:
            continue
        visited.add(url)
        try:
            html = fetch(url)
        except Exception:
            continue
        text = _html_to_text(html)
        if text.strip():
            pages.append((url, text))
        for link in _extract_links(html, url):
            if link not in visited and (not same_domain or urlparse(link).netloc == domain):
                queue.append(link)
    return pages


def _chunks_for(
    text: str, source: str, cfg: Settings, base_metadata: dict
) -> list[Chunk]:
    text = text.strip()
    if not text:
        return []
    return chunk_document(
        text,
        source=source,
        chunk_tokens=cfg.chunk_tokens,
        chunk_overlap=cfg.chunk_overlap,
        base_metadata=base_metadata,
    )


def load_chunks(
    paths: Iterable[str | Path],
    *,
    urls: Iterable[str] | None = None,
    crawl_urls: Iterable[str] | None = None,
    crawl_max_pages: int = 20,
    cfg: Settings = settings,
    tenant_id: str | None = None,
) -> list[Chunk]:
    """Read files, single URLs, and crawled sites into chunks, tagged with tenant."""
    common = {"tenant_id": tenant_id} if tenant_id else {}
    chunks: list[Chunk] = []

    for path in discover_files(paths):
        loader = LOADERS[path.suffix.lower()]
        meta = {**common, "filename": path.name, "suffix": path.suffix.lower()}
        chunks.extend(_chunks_for(loader(path), str(path), cfg, meta))

    for url in urls or []:
        meta = {**common, "source_type": "url"}
        chunks.extend(_chunks_for(fetch_url(url), url, cfg, meta))

    for start_url in crawl_urls or []:
        for page_url, text in crawl_site(start_url, max_pages=crawl_max_pages):
            meta = {**common, "source_type": "crawl", "crawl_root": start_url}
            chunks.extend(_chunks_for(text, page_url, cfg, meta))

    return chunks


def build_store(
    paths: Iterable[str | Path],
    *,
    urls: Iterable[str] | None = None,
    crawl_urls: Iterable[str] | None = None,
    crawl_max_pages: int = 20,
    cfg: Settings = settings,
    embedder: Embedder | None = None,
    store: VectorStore | None = None,
    tenant_id: str | None = None,
) -> VectorStore:
    """Load, embed, and index files/URLs/crawls into a (new or existing) store."""
    embedder = embedder or build_embedder(cfg)
    # `is not None` (not `or`): an empty store is falsy, and truthiness on the
    # pgvector backend would query a table that may not exist yet.
    store = store if store is not None else VectorStore()

    chunks = load_chunks(
        paths,
        urls=urls,
        crawl_urls=crawl_urls,
        crawl_max_pages=crawl_max_pages,
        cfg=cfg,
        tenant_id=tenant_id,
    )
    if not chunks:
        raise ValueError(
            f"No ingestible content found under: {list(paths)} / urls={list(urls or [])}"
        )

    vectors = embedder.embed([c.text for c in chunks], input_type="document")
    store.add(chunks, vectors, tenant_id=tenant_id)
    return store
